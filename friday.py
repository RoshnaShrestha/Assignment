import openpyxl

def create_messy_dataset(file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sales_Data"  # Fixed invalid title

    # Adding messy data
    sheet.append(["Data", "Product", "Quantity", "Price"])
    sheet.append(["34/45/67", "Laptop", 3, 1200.50])
    sheet.append(["", "Mouse", "", 25.99])
    sheet.append(["34/45/67", "Laptop", 3, 1200.50])
    sheet.append(["34/45/67", "Keyboard", 2, "None"])  # "None" as a string
    workbook.save(file_path)
    print("Messy dataset created.")

def clean_dataset(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    cleaned_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Fixed type: `value_only` to `values_only`
        # Replace "None" (string) with None and skip rows with empty or missing values
        row = tuple(cell if cell != "None" else None for cell in row)
        if None in row or "" in row:  # Skip rows with empty values
            continue
        if row not in cleaned_data:  # Avoid duplicates
            cleaned_data.append(row)

    # Clear the sheet content
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.value = None

    # Add headers and cleaned data
    headers = ["Data", "Product", "Quantity", "Price"]
    sheet.append(headers)
    for row in cleaned_data:
        sheet.append(row)

    workbook.save(file_path)
    print("Cleaned dataset saved.")

file_path = "messy_sales_data.xlsx"
create_messy_dataset(file_path)
clean_dataset(file_path)



