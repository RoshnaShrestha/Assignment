"""
Application Task: Student Performance Tracker
Objective:
Create 
a Python application that reads student data from a CSV file, calculates averages, and writes the results to an Excel file. Use the collections module to analyze the data.

Scenario:
You are tasked with developing a small application for a school to track and analyze students' performance in multiple subjects.

Requirements:
Input File: Create a CSV file named student_scores.csv with the following content:

Name,Subject,Score
Alice,Math,85
Bob,Math,90
Alice,Science,95
Bob,Science,80
Charlie,Math,70
Charlie,Science,75


Features:
Read Data: Read the CSV file and store the data in a suitable collection (e.g., a list of dictionaries).
Analyze Data:
Calculate the average score for each student.
Identify the subject with the highest average score.
Write Results:
Save the calculated averages and highest scoring subject to an Excel file (results.xlsx).
The Excel file should have two sheets:
Student Averages: Contains each student's average score.
Subject Averages: Contains the average score for each subject.

Suggested Output:
Excel File Structure:
Sheet 1: Student Averages:
Copy code
Name,Average Score
Alice,90
Bob,85
Charlie,72.5


Sheet 2: Subject Averages:
javascript
Copy code
Subject,Average Score
Math,81.67
Science,83.33
----------------------------------------------------------------------------------------------------------

import csv
import numpy as np
import openpyxl

student_file_path = "student_scores.csv"
summary_file_path = "results.xlsx"

def create_csv():
    # Create a CSV file with headers
    headers = ["Name", "Subject", "Score"]
    with open(student_file_path, mode="w", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(headers)
    print("Student information file created.")

def add_record(name, subject, score):
    # Append a new record to the CSV file
    with open(student_file_path, mode="a", newline="") as file:  
        writer = csv.writer(file)
        writer.writerow([name, subject, score])  
    print(f"Record added: {name}, {subject}, {score}")

def calculate_averages():
    # Analyze the data in the CSV file
    total_students = 0
    student_scores = {}
    subject_scores = {}

    with open(student_file_path, mode="r") as file:
        reader = csv.DictReader(file)
        for row in reader:
            total_students += 1
            name = row["Name"]
            subject = row["Subject"]
            score = float(row["Score"])

            # Group scores by student
            if name in student_scores:
                student_scores[name].append(score)
            else:
                student_scores[name] = [score]

            # Group scores by subject
            if subject in subject_scores:
                subject_scores[subject].append(score)
            else:
                subject_scores[subject] = [score]

    # Calculate average scores for students and subjects
    student_averages = {name: np.mean(scores) for name, scores in student_scores.items()}
    subject_averages = {subject: np.mean(scores) for subject, scores in subject_scores.items()}

    return total_students, student_averages, subject_averages

def save_to_excel(student_averages, subject_averages):
    # Create an Excel file with two sheets: Student Averages and Subject Averages
    workbook = openpyxl.Workbook()

    # Sheet 1: Student Averages
    sheet1 = workbook.active
    sheet1.title = "Student Averages"
    sheet1.append(["Name", "Average Score"])
    for name, avg_score in student_averages.items():
        sheet1.append([name, round(avg_score, 2)])

    # Sheet 2: Subject Averages
    sheet2 = workbook.create_sheet(title="Subject Averages")
    sheet2.append(["Subject", "Average Score"])
    for subject, avg_score in subject_averages.items():
        sheet2.append([subject, round(avg_score, 2)])

    workbook.save(summary_file_path)
    print(f"Results saved to Excel file: {summary_file_path}")

def create_messy_dataset(file_path):
    # Create an Excel file with messy data
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Student_Data"

    # Append rows of data
    sheet.append(["Name", "Subject", "Score"])
    sheet.append(["Alice", "Math", 85])
    sheet.append(["Bob", "Math", 90])
    sheet.append(["Alice", "Science", 95])
    sheet.append(["Bob", "Science", 80])
    sheet.append(["Charlie", "Math", 70])
    sheet.append(["Charlie", "Science", 75])

    workbook.save(file_path)
    print(f"Messy dataset created at: {file_path}")

# Main Execution
create_csv()

# Add records to the CSV file
add_record("Alice", "Math", 85)
add_record("Bob", "Math", 90)
add_record("Alice", "Science", 95)
add_record("Bob", "Science", 80)
add_record("Charlie", "Math", 70)
add_record("Charlie", "Science", 75)

# Analyze data and save results
total_students, student_averages, subject_averages = calculate_averages()

print(f"Total Students: {total_students}")
print("Average Scores:")
for name, avg_score in student_averages.items():
    print(f"{name}: {avg_score:.2f}")
print("Subject Averages:")
for subject, avg_score in subject_averages.items():
    print(f"{subject}: {avg_score:.2f}")

# Save results to Excel
save_to_excel(student_averages, subject_averages)

# Create a messy dataset as an Excel file
create_messy_dataset("messy_data.xlsx")
"""


import csv
from collections import defaultdict
from openpyxl import Workbook


